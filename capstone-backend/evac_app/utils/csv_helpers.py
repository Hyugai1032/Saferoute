"""
csv_helpers.py - FINAL VERSION with proper coordinate handling
Place this in your utils folder
"""

import re
from openpyxl import load_workbook
import csv
from io import StringIO

def normalize_header(h: str) -> str:
    if h is None:
        return ""
    h = str(h)

    # remove BOM + quotes + normalize whitespace/newlines
    h = h.replace("\ufeff", "")
    h = h.replace("\n", " ").replace("\r", " ")
    h = h.replace('"', "").replace("'", "")
    h = h.strip().lower()
    h = re.sub(r"\s+", " ", h)

    return h



def dms_to_decimal(dms_str):
    """Convert DMS coordinates to decimal format"""
    if not dms_str or not isinstance(dms_str, str):
        return None
    dms_str = str(dms_str).strip()
    
    # Try numeric decimal first
    try:
        return float(dms_str)
    except ValueError:
        pass
    
    # Handle "Lat: 13 08.278" format (common in your Excel)
    if "lat" in dms_str.lower() or "long" in dms_str.lower():
        # Extract just the number part
        match = re.search(r'([-+]?\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)', dms_str)
        if match:
            degrees = float(match.group(1))
            minutes = float(match.group(2))
            return degrees + (minutes / 60.0)
    
    # DMS regex pattern for full format
    pattern = r"""(?P<deg>-?\d+)[°\s]+
                  (?P<min>\d+)[\'\s]+
                  (?P<sec>\d+(?:\.\d+)?)[\"\s]*
                  (?P<dir>[NSEW])?"""
    match = re.match(pattern, dms_str, re.VERBOSE | re.IGNORECASE)
    
    if match:
        deg = float(match.group("deg"))
        minutes = float(match.group("min"))
        seconds = float(match.group("sec"))
        direction = match.group("dir")
        
        decimal = deg + (minutes / 60.0) + (seconds / 3600.0)
        
        if direction and direction.upper() in ["S", "W"]:
            decimal = -decimal
        
        return decimal
    
    # Fallback: try splitting by spaces
    try:
        parts = re.split(r'[,\s]+', dms_str)
        if len(parts) >= 1:
            return float(parts[0])
    except Exception:
        pass
    
    return None


def read_csv_rows(file_obj):
    if hasattr(file_obj, 'read'):
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig', errors='ignore')
    else:
        content = str(file_obj)

    f = StringIO(content)
    reader = csv.DictReader(f)

    if reader.fieldnames:
        reader.fieldnames = [normalize_header(h) for h in reader.fieldnames]

    rows = []
    for row in reader:
        rows.append({normalize_header(k): v for k, v in row.items()})
    return rows



def read_xlsx_rows(file):
    import openpyxl

    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb.active

    all_rows = list(sheet.iter_rows(values_only=True))

    # ---- find the "parent header" row (Location, Name of Facility, Coordinates, Capacity, etc.)
    header_row_idx = None
    parent = None
    for idx, row in enumerate(all_rows[:30], start=1):
        row_lower = [normalize_header(cell) for cell in row]
        has_location = any("location" == c for c in row_lower)
        has_facility = any("name of facility" in c or "facility" in c for c in row_lower)
        has_coords = any("coordinates" in c for c in row_lower)
        if has_location and has_facility and has_coords:
            header_row_idx = idx
            parent = row
            break

    if not header_row_idx:
        print("❌ ERROR: Could not find header row!")
        return []

    # ---- child header is the NEXT row (Province/Municipality/Barangay + Families/Individuals)
    child = all_rows[header_row_idx] if header_row_idx < len(all_rows) else None

    # ---- build column_map using parent+child headers
    column_map = {}
    for col_idx in range(len(parent)):
        p = normalize_header(parent[col_idx])
        c = normalize_header(child[col_idx]) if child and col_idx < len(child) else ""

        # Your sheet structure:
        # parent: Location -> child: Province/Municipality/Barangay
        if p == "location" and c in {"province", "municipality", "barangay"}:
            key = c

        # parent: Capacity -> child: Families/Individuals (we want keys EXACTLY "families"/"individuals")
        elif p == "capacity" and c in {"families", "individuals"}:
            key = c

        # if child header exists, prefer it (for clean keys)
        elif c:
            key = c

        # else use parent header
        elif p:
            key = p

        else:
            key = f"col_{col_idx}"

        column_map[col_idx] = key

    # ---- process data rows (actual data starts after the child header row)
    data_start = header_row_idx + 1  # because header_row_idx is 1-based, and we used parent+child
    rows = []

    current_province = None
    current_municipality = None
    current_barangay = None

    last_facility_row = None
    pending_lat = None

    for row_idx in range(data_start, len(all_rows)):
        row_values = all_rows[row_idx]

        # skip empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        row_dict = {}
        for col_idx, value in enumerate(row_values):
            if col_idx not in column_map:
                continue
            if value is None:
                continue
            s = str(value).strip()
            if s == "":
                continue
            row_dict[column_map[col_idx]] = value

        # carry forward merged cells for province/municipality/barangay
        prov = str(row_dict.get("province", "")).strip()
        mun = str(row_dict.get("municipality", "")).strip()
        brgy = str(row_dict.get("barangay", "")).strip()

        if prov and "province" not in prov.lower():
            current_province = prov
        elif current_province:
            row_dict["province"] = current_province

        if mun and "municipality" not in mun.lower():
            current_municipality = mun
        elif current_municipality:
            row_dict["municipality"] = current_municipality

        if brgy and "barangay" not in brgy.lower():
            current_barangay = brgy
        elif current_barangay:
            row_dict["barangay"] = current_barangay

        # standardize facility name into "name of facility"
        facility = None
        for k in ["name of facility", "facility name", "facility", "name", "evacuation center"]:
            if k in row_dict:
                facility = str(row_dict[k]).strip()
                if facility:
                    row_dict["name of facility"] = facility
                    break

        # standardize coordinates into ONE key
        coord_text = str(
            row_dict.get("coordinates (latitude and longitude)")
            or row_dict.get("coordinates")
            or ""
        ).strip()

        # ✅ IMPORTANT: In your file, the facility row contains "Lat: ...", the next row contains "Long: ..."
        # So we DO NOT skip the facility row.
        if coord_text:
            if "lat" in coord_text.lower():
                pending_lat = coord_text
                # ensure the key is consistent
                row_dict["coordinates (latitude and longitude)"] = coord_text

            elif "long" in coord_text.lower():
                # longitude-only row: attach to last facility row
                if pending_lat and last_facility_row is not None:
                    last_facility_row["coordinates (latitude and longitude)"] = f"{pending_lat}, {coord_text}"
                pending_lat = None
                continue  # do not add the longitude-only row as a record

        # Only save rows that actually represent a facility
        if facility:
            rows.append(row_dict)
            last_facility_row = rows[-1]

    return rows